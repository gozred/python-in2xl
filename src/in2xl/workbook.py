# in2xl.py
from __future__ import annotations
from openpyxl.utils.dataframe import dataframe_to_rows
from xlsxwriter.utility import xl_col_to_name as xl_name
import ruamel.std.zipfile as zipfile
from lxml import etree
from datetime import datetime
from typing import Union
import shutil
import numbers
import decimal
import pandas as pd
import os
import re

XMAIN = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
XREL = '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}'
XCHART = '{http://schemas.openxmlformats.org/drawingml/2006/chart}'
XDSGN = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac"


class Workbook:

    """
    The `Workbook` class represents an Excel workbook.

    Attributes:
        xmain (str): A string representing the main namespace.
        xrel (str): A string representing the relationship namespace.
        xchart (str): A string representing the chart namespace.
        xdsgn (str): A string representing the drawing namespace.
        temp (str): A string representing the temporary path where the Excel workbook is copied.
        wb (Element): An Element object representing the workbook.xml file.
        wb_dict (dict): A dictionary of worksheet names and their corresponding filenames.
        sheetnames (list):  A list of the names of all sheetnames in the Excel workbook.
        wb_id_dict (dict): A dictionary of worksheet names and their corresponding sheet Ids.
        content (list): A list of all files and directories in the zip archive.
        chart_dict (dict): A dictionary of worksheet names and the corresponding chart filenames.
        wb_state (dict): A dictionary of worksheet names and their corresponding states (visible, hidden, etc.).

    Methods:
        __init__(self): Initializes a new instance of the `Workbook` class.
        __worksheets(self): Extracts worksheet information from a given Excel file.
        load_workbook(self, path: str = None) -> Worksheets: Reads an Excel workbook from the specified file path and returns an instance of Worksheets.

    """
    def __new__(cls, path=None):

        if path is not None:
            cls.__init__(cls)

            return object.__new__(cls).load_workbook(path)

        return object.__new__(cls)

    def __init__(self):

        self.xmain = XMAIN
        self.xrel = XREL
        self.xchart = XCHART
        self.xdsgn = XDSGN

        self.temp = None
        self.wb = None
        self.wb_dict = self.wb_id_dict = self.wb = self.content = self.chart_dict = self.wb_state = None

    def __worksheets(self):

        """
        This method extracts worksheet information from a given Excel file.

        Parameters:
            None

        Returns:
            wb_dict (dict): A dictionary of worksheet names and their corresponding filenames.
            wb_id_dict (dict): A dictionary of worksheet names and their corresponding sheet Ids.
            wb (Element): An Element object representing the workbook.xml file.
            content (list): A list of all files and directories in the zip archive.
            chart_dict (dict): A dictionary of worksheet names and the corresponding chart filenames.
            wb_state (dict): A dictionary of worksheet names and their corresponding states (visible, hidden, etc.).

        """

        with zipfile.ZipFile(self.temp, mode="r") as myzip:
            with myzip.open('xl/workbook.xml') as myfile:
                wb = etree.fromstring(myfile.read())
                content = myzip.namelist()

        wb_dict = {}
        wb_id_dict = {}
        wb_state = {}
        chart_dict = {}

        # Extract information from the workbook.xml file

        for sheets in wb.iter(f'{self.xmain}sheets'):
            for i in sheets:
                wb_dict[i.attrib['name']] = 'sheet{no}.xml'.format(no=i.attrib[f'{self.xrel}id'][3:])
                wb_id_dict[i.attrib['name']] = i.attrib['sheetId']
                if 'state' in i.attrib:
                    wb_state[i.attrib['name']] = i.attrib['state']
                else:
                    wb_state[i.attrib['name']] = 'visible'

        # Check for charts
        charts = [i for i in content if i.startswith('xl/charts/chart')]

        if charts:
            # Extract chart information
            for c in charts:
                with zipfile.ZipFile(self.temp, mode="r") as myzip:
                    with myzip.open(c) as myfile:
                        chart = etree.fromstring(myfile.read())

                __a__ = []
                for items in chart.iter(f'{self.xchart}f'):
                    __a__.append(items.text)

                for i in wb_dict.keys():
                    if i.__contains__("'"):
                        j = i.replace("'", "''")
                    else:
                        j = i
                    if any(j in word for word in __a__):
                        if i in chart_dict.keys():
                            chart_dict[i] += [c]
                        else:
                            chart_dict[i] = [c]

        self.sheetnames = list(wb_dict)

        return wb_dict, wb_id_dict, wb, content, chart_dict, wb_state

    def load_workbook(self, path: str = None) -> Worksheets:

        """
        Reads an Excel workbook from the specified file path and returns an instance of Worksheets.

        Args:
            path (str): The file path of the Excel workbook to read.

        Returns:
            Worksheets: An instance of the Worksheets class that contains the data extracted from the workbook.

        Raises:
            ValueError: If the `path` argument is None.

        """

        if path is None:
            raise ValueError('Templatepath is missing')

        tpath = os.path.dirname(path)
        dtime = datetime.now().strftime("%Y%m%d")
        tfile = f"~{dtime}_{os.path.basename(path)}.zip"

        self.temp = os.path.join(tpath, tfile)

        shutil.copy(path, self.temp)

        self.wb_dict, self.wb_id_dict, self.wb, self.content, self.chart_dict, self.wb_state = self.__worksheets()

        return Worksheets(parent=self)


class Worksheets():

    """
    Represents a collection of worksheets extracted from an Excel workbook.

    Args:
        parent (object): The parent object.
        key (str): The key for the current worksheet.

    Attributes:
        temp (str): The path of the temporary file created to hold the Excel workbook data.
        xmain (str): The XML namespace for main XML elements.
        xdsgn (str): The XML namespace for design XML elements.
        wb_dict (dict): A dictionary mapping worksheet names to their corresponding XML file names.
        sheetnames (list):  A list of the names of all sheetnames in the Excel workbook.
        wb_id_dict (dict): A dictionary mapping worksheet names to their corresponding IDs.
        wb (etree._Element): The XML tree representing the workbook.
        wb_state (dict): A dictionary mapping worksheet names to their corresponding visibility states.
        content (list): A list of the names of all files in the Excel workbook.
        stree (None): A reference to the worksheet XML tree.
        key (str): The name of the current worksheet.
        _state (str): The visibility state of the current worksheet.
        sheet (str): The name of the XML file that represents the current worksheet.
        tree (None): A reference to the XML tree for the current worksheet.
        never (bool): A flag that indicates whether the current worksheet has never been accessed.
        check (tuple): A tuple of numeric types to check against for float values.

    Methods:
        __init__(self, parent=None, key=None): Initializes a new instance of the Worksheets class.
        __getitem__(self, key): Retrieves a worksheet by name from the Excel workbook.
        state(self) -> str:  Returns/Sets the state of the worksheets object.
        __change_dim(self, xml, row, column): Changes the dimensions of the specified cell range.
        __change_xml(self, xml, row, column, value): Insert data into an XML sheet.
        __change_strxml(self, value): Find, add or changes the value of a specified XML tag within the string table element
        __create_SubEl(self, main, tag, attrib={}, text=None): Creates a new sub-element with the given tag and attributes under the specified main element.
        __clean_formula(self): Removes any child elements with tag 'v' under each 'f' element in the XML tree of the class instance.
        __get_strxml(self): Retrieves and parses the XML content from the 'xl/sharedStrings.xml' file in the Excel workbook.
        __get_xml(self): Reads the XML data for the current worksheet from the temporary file and stores it in `self.tree`.
        __write_state(self, value): Update the state attribute of a sheet in the workbook.
        __write_xml(self): This method writes the current XML tree to the corresponding worksheet file within the Excel workbook file.
        __write_strxml(self): Write the shared strings XML to the temporary zip file.
        close(self) -> None: Close the workbook by removing the temporary file.
        insert(self, data: Union(str, int, float, pd.DataFrame), row: int = 1, column: int = 1, axis: int = 0, header: bool = True, index: bool = False ) -> None: Insert data into the worksheet. Convert the input data into an array and pass it to the XML converter.
        save(self, path: str = None) -> None: Saves the converted Excel file to the specified path

    """

    def __init__(self, parent=None, key=None):
        self.temp = parent.temp
        self.xmain = parent.xmain
        self.xdsgn = parent.xdsgn
        self.wb_dict = parent.wb_dict
        self.wb_id_dict = parent.wb_id_dict
        self.wb = parent.wb
        self.wb_state = parent.wb_state
        self.content = parent.content
        self.sheetnames = parent.sheetnames
        self.stree = None
        self.key = key
        if key is not None:
            self._state = self.wb_state[self.key]
            self.sheet = self.wb_dict[self.key]
        self.tree = None
        self.never = False
        self.check = (numbers.Real, decimal.Decimal)

    def __getitem__(self, key):
        if self.temp is None:
            raise ValueError('please choose a workbook first')

        if key not in self.wb_dict.keys():
            sl = list(self.wb_dict.keys())
            raise KeyError(f"the sheet [{key}] seems to be not included in this Excel workbook, possible sheets: {sl}")

        return self.__class__(parent=self, key=key)

    @property
    def state(self) -> str:
        """
        Returns the current state of the worksheets object.

        Returns:
            str: The current state of the worksheets object.
        """
        return self._state

    @state.setter
    def state(self, value: int) -> Worksheets:
        """
        Sets the state of the worksheets object to the specified value.

        Args:
            value (int): The new state value to set. Must be an integer within the range [0, 2].

        Returns:
            Worksheets: The updated Worksheets object.

        Raises:
            ValueError: If the input value is not an integer or is outside the allowed range.
        """

        if (not isinstance(value, int)) or (abs(int(value)) > 2):
            raise ValueError(f'Input is outside of the parameters. (0-2 expected, {value} received)')

        if value == 0:
            self._state = 'visible'
        elif value == 1:
            self._state = 'hidden'
        elif value == 2:
            self._state = 'veryHidden'

        self.__write_state(self._state)

        return self

    def __change_dim(self, xml, row, column):

        """
        Changes the dimensions of the specified cell range.

        Args:
            xml (Element): The XML element to modify.
            row (int): The row number of the cell range.
            column (str): The column letter of the cell range.

        Returns:
            Worksheets: The updated Worksheets object.
        """

        dimnav = xml.find(".//x:dimension", namespaces={'x': f'{self.xmain}'.strip('{}')})
        match = re.match(r"([A-Z]+)(\d+)\:{1}([A-Z]+)(\d+)", dimnav.attrib['ref'], re.I)
        fl = sum([(ord(match.groups()[0].upper()[-i-1])-64)*26**i for i in range(len(match.groups()[0]))])
        fn = int(match.groups()[1])
        sl = sum([(ord(match.groups()[2].upper()[-i-1])-64)*26**i for i in range(len(match.groups()[2]))])
        sn = int(match.groups()[3])

        cl = sum([(ord(column.upper()[-i-1])-64)*26**i for i in range(len(column))])

        if fl < cl:
            pass
        elif fl == cl:
            if fn <= row:
                pass
            else:
                fn = row
        else:
            fl = cl

        if sl > cl:
            pass
        elif sl == cl:
            if sn >= row:
                pass
            else:
                sn = row
        else:
            sl = cl

        att = f'{xl_name(fl-1)}{fn}:{xl_name(sl-1)}{sn}'

        dimnav.attrib['ref'] = att

        return self

    def __change_xml(self, xml, row, column, value):
        """
        Insert data into an XML sheet.

        Parameters:
            xml (ElementTree.Element): The XML sheet to be modified.
            row (int): The row number where the value will be inserted.
            column (str): The column letter where the value will be inserted.
            value (str or int): The value to be inserted.

        Returns:
            self: The modified XML sheet.
        """

        ws_row = xml.find(f"./{self.xmain}sheetData/{self.xmain}row/[@r='{row}']")

        if ws_row is None:

            a_row = xml.findall(".//x:row", namespaces={'x': f'{self.xmain}'.strip('{}')})

            if not a_row:

                ws_main = xml.find(f"./{self.xmain}sheetData")
                self.__create_SubEl(ws_main, f'{self.xmain}row', attrib={'r': str(row),
                                                                         'spans': '1:1',
                                                                         etree.QName(self.xdsgn, 'dyDescent'): '0.25'
                                                                         }
                                    )
                dimnav = xml.find(".//x:dimension", namespaces={'x': f'{self.xmain}'.strip('{}')})
                dimnav.attrib['ref'] = f'{column}{row}:{column}{row}'

                self.__change_xml(xml, row, column, value)

                return self

            srlist = [int(x.attrib['r']) for x in a_row]
            # for i in a_row:
            #    srlist.append(int(i.attrib['r']))
            min_n = min(srlist, key=lambda x: (abs(x - row), x))
            add_row = xml.find(f".//x:row[@r='{min_n}']", namespaces={'x': f'{self.xmain}'.strip('{}')})
            new_row = etree.Element(f'{self.xmain}row')
            for i in add_row.attrib.keys():
                new_row.attrib[i] = add_row.attrib[i]
            new_row.attrib['r'] = str(row)
            if min_n > row:
                add_row.addprevious(new_row)
            else:
                add_row.addnext(new_row)

            self.__change_dim(xml, row, column)

            self.__change_xml(xml, row, column, value)

            return self

        ws_column = ws_row.find(f"./{self.xmain}c/[@r='{column}{row}']")

        if ws_column is None:

            oc = sum([(ord(column.upper()[-i-1])-64)*26**i for i in range(len(column))])
            a_row = xml.findall(".//x:row", namespaces={'x': f'{self.xmain}'.strip('{}')})
            rownav = xml.find(f".//x:row[@r='{row}']", namespaces={'x': f'{self.xmain}'.strip('{}')})
            allrow = rownav.findall(".//x:c", namespaces={'x': f'{self.xmain}'.strip('{}')})

            if not allrow:

                istext = ''

                if isinstance(value, str):
                    value = self.__change_strxml(value)
                    istext = ' t="s"'
                    self.never = False

                ws_row.append(etree.XML(f'<c r="{column}{row}"{istext}><v>{value}</v></c>'))

                return self

            c_vault = []

            for i in allrow:
                match = re.match(r"(^[A-Z]+)", i.attrib['r'], re.I)
                c_vault.append(sum([(ord(match.group().upper()[-i-1])-64)*26**i for i in range(len(match.group()))]))

            c_min = min(c_vault, key=lambda x: (abs(x - oc), x))

            add_c = rownav.find(f".//x:c[@r='{xl_name(c_min-1)}{row}']", namespaces={'x': f'{self.xmain}'.strip('{}')})
            new_c = etree.Element(f'{self.xmain}c')
            new_c.attrib['r'] = f'{xl_name(oc-1)}{row}'

            if c_min > oc:
                add_c.addprevious(new_c)
            else:
                add_c.addnext(new_c)

            self.__change_dim(xml, row, column)

            self.__change_xml(xml, row, column, value)

            return self

        istext = ''

        if isinstance(value, str):
            value = self.__change_strxml(value)
            istext = ' t="s"'
            self.never = False

        ws_value = ws_column.find(f"./{self.xmain}v")

        if (istext != '') and ('t' not in ws_column.attrib):
            ws_column.set('t', "s")
        elif (istext == '') and ('t' in ws_column.attrib):
            ws_column.attrib.pop('t')

        if ws_value is None:
            ws_column.append(etree.XML(f'<v>{value}</v>'))
        elif ws_value is not None:
            ws_value.text = str(value)

        return self

    def __change_strxml(self, value):
        """
        Find, add or changes the value of a specified XML tag within the string table element.

        Args:
            value (str): The value to change the XML tag to.

        Returns:
            int: The index number of the updated XML tag within the string table element.

        """

        if self.stree.xpath(f'.//x:t[text()="{value}"]', namespaces={'x': f'{self.xmain}'.strip('{}')}):
            _full_list = self.stree.xpath('.//x:t', namespaces={'x': f'{self.xmain}'.strip('{}')})
            _count = 0
            for _data in _full_list:
                if _data.text == value:
                    _number = _count
                    self.stree.attrib['count'] = str(int(self.stree.attrib['count']) + 1)
                else:
                    _count = _count + 1
        else:
            root = self.stree.getroottree().getroot()
            self.__create_SubEl(root, f'{self.xmain}si')
            sub_si = self.stree.xpath(".//x:si", namespaces={'x': f'{self.xmain}'.strip('{}')})[-1]
            self.__create_SubEl(sub_si, f'{self.xmain}t', text=value)
            self.stree.attrib['count'] = str(int(self.stree.attrib['count']) + 1)
            self.stree.attrib['uniqueCount'] = str(int(self.stree.attrib['uniqueCount']) + 1)
            _number = int(self.stree.attrib['uniqueCount']) - 1

        return _number

    def __create_SubEl(self, main, tag, attrib={}, text=None):

        """
        Creates a new sub-element with the given tag and attributes under the specified main element.

        Args:
            main: The main element under which the sub-element will be created.
            tag: The tag of the sub-element to be created.
            attrib: A dictionary of attributes for the sub-element (optional).
            text: The text content of the sub-element (optional).

        Returns:
            The instance of the class.

        """

        node = etree.SubElement(main, tag, attrib)
        node.text = text

        return self

    def __clean_formula(self):

        """
        Removes any child elements with tag 'v' under each 'f' element in the XML tree of the class instance.

        """

        form_clean = self.tree.findall('.//x:f', namespaces={'x': f'{self.xmain}'.strip('{}')})
        for e in form_clean:
            parent = e.getparent()
            for c in parent.getchildren():
                if c.tag == '{s}v'.format(s=self.xmain):
                    parent.remove(c)

    def __get_strxml(self):

        """
        Retrieves and parses the XML content from the 'xl/sharedStrings.xml' file in the Excel workbook.

        Returns:
            The instance of the class.

        Raises:
            KeyError: If 'xl/sharedStrings.xml' file is not found in the Excel workbook.

        """

        try:
            with zipfile.ZipFile(self.temp, mode="r") as myzip:
                with myzip.open('xl/sharedStrings.xml') as myfile:
                    self.stree = etree.fromstring(myfile.read())

            zipfile.delete_from_zip_file(self.temp, file_names='xl/sharedStrings.xml')

        except KeyError:
            self.stree = etree.fromstring(b'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="0" uniqueCount="0"/>')
            self.never = True

            with zipfile.ZipFile(self.temp, mode="r") as myzip:
                with myzip.open('[Content_Types].xml') as myfile:
                    cttree = etree.fromstring(myfile.read())

            xct = '{http://schemas.openxmlformats.org/package/2006/content-types}'

            ct_check = cttree.find(f"./{xct}Override/[@PartName='/xl/sharedStrings.xml']")
            if ct_check is None:
                zipfile.delete_from_zip_file(self.temp, file_names='[Content_Types].xml')

                pre_ct = cttree.find(f"./{xct}Override/[@PartName='/xl/styles.xml']")
                new_ct = etree.Element(f'{xct}Override')
                new_ct.attrib['ContentType'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"
                new_ct.attrib['PartName'] = "/xl/sharedStrings.xml"

                pre_ct.addnext(new_ct)

                with zipfile.ZipFile(self.temp, mode="a") as myzip:
                    with myzip.open('[Content_Types].xml', 'w') as myfile:
                        myfile.write(etree.tostring(cttree))

            with zipfile.ZipFile(self.temp, mode="r") as myzip:
                with myzip.open('xl/_rels/workbook.xml.rels') as myfile:
                    retree = etree.fromstring(myfile.read())

            xre = "{http://schemas.openxmlformats.org/package/2006/relationships}"

            re_check = retree.find(f"./{xre}Relationship/[@Target='sharedStrings.xml']")
            if re_check is None:

                zipfile.delete_from_zip_file(self.temp, file_names='xl/_rels/workbook.xml.rels')

                all_re = retree.findall('.//x:Relationship', namespaces={'x': f'{xre}'.strip('{}')})

                self.__create_SubEl(retree, f'{xre}Relationship',
                                    attrib={'Target': "sharedStrings.xml",
                                            'Type': "http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings",
                                            'Id': f"rId{len(all_re)+1}"
                                            }
                                    )

                with zipfile.ZipFile(self.temp, mode="a") as myzip:
                    with myzip.open('xl/_rels/workbook.xml.rels', 'w') as myfile:
                        myfile.write(etree.tostring(retree))

        return self

    def __get_xml(self):

        """
        Reads the XML data for the current worksheet from the temporary file and stores it in `self.tree`.
        Then, it deletes the worksheet file from the temporary ZIP archive.

        Returns:
            self : Returns the instance of the `Excel` class after reading and storing the worksheet XML data.

        """

        with zipfile.ZipFile(self.temp, mode="r") as myzip:
            with myzip.open(f'xl/worksheets/{self.sheet}') as myfile:
                self.tree = etree.fromstring(myfile.read())

        zipfile.delete_from_zip_file(self.temp, file_names=f'xl/worksheets/{self.sheet}')

        return self

    def __write_state(self, value):

        """
        Update the state attribute of a sheet in the workbook.

        Args:
            value (str): The new state value for the sheet.

        Returns:
            self: The updated instance of the class.

        Raises:
            IndexError: If the sheet with the given name does not exist in the workbook.

        """

        # If the current state is the same as the new value, return the instance without modifying anything.
        if self.wb_state[self.key] == value:
            return self

        # Delete the workbook.xml file from the zip file.
        zipfile.delete_from_zip_file(self.temp, file_names='xl/workbook.xml')

        # Find the sheet node in the workbook with the given name.
        snode = self.wb.xpath(f'.//x:sheet[@name="{self.key}"]', namespaces={'x': f'{self.xmain}'.strip('{}')})[0]

        if value == 'visible':
            snode.attrib.pop('state')
        else:
            snode.attrib['state'] = value

        # Write the updated workbook xml to the zip file.
        with zipfile.ZipFile(self.temp, mode="a") as myzip:
            with myzip.open('xl/workbook.xml', 'w') as myfile:
                myfile.write(etree.tostring(self.wb))

        with zipfile.ZipFile(self.temp, mode="r") as myzip:
            with myzip.open('xl/workbook.xml') as myfile:
                self.wb = etree.fromstring(myfile.read())

        wb_state = {}

        for sheets in self.wb.iter(f'{self.xmain}sheets'):
            for i in sheets:
                if 'state' in i.attrib:
                    wb_state[i.attrib['name']] = i.attrib['state']
                else:
                    wb_state[i.attrib['name']] = 'visible'

        self.wb_state = wb_state

        return self

    def __write_xml(self):

        """
        This method writes the current XML tree to the corresponding worksheet file within the Excel workbook file.

        Returns:
            Workbook: instance of the Workbook class.

        """

        with zipfile.ZipFile(self.temp, mode="a") as myzip:
            with myzip.open(f'xl/worksheets/{self.sheet}', 'w') as myfile:
                myfile.write(etree.tostring(self.tree))

        return self

    def __write_strxml(self):

        """
        Write the shared strings XML to the temporary zip file.

        Returns:
            Workbook: instance of the Workbook class.

        """

        if self.never:
            return self

        with zipfile.ZipFile(self.temp, mode="a") as myzip:
            with myzip.open('xl/sharedStrings.xml', 'w') as myfile:
                myfile.write(etree.tostring(self.stree))

        return self

    def close(self) -> None:
        """
        Close the workbook by removing the temporary file.

        """

        return os.remove(self.temp)

    def insert(self,
               data: Union(str, int, float, pd.DataFrame),
               row: int = 1,
               column: int = 1,
               axis: int = 0,
               header: bool = True,
               index: bool = False
               ) -> None:
        """
        Insert data into the worksheet. Convert the input data into an array and pass it to the XML converter.

        Args:
            data (Union[str, int, float, pd.DataFrame]): Data to be inserted.
            row (int, optional): Row number where the data is to be inserted. Defaults to 1.
            column (int, optional): Column number where the data is to be inserted. Defaults to 1.
            axis (int, optional): 0 to insert data row-wise and 1 to insert data column-wise. Defaults to 0.
            header (bool, optional): True to include headers in the data, False otherwise. Defaults to True.
            index (bool, optional): True to include index in the data, False otherwise. Defaults to False.

        """

        column = column - 1

        self.__get_xml()
        self.__get_strxml()

        if isinstance(data, pd.core.frame.DataFrame):

            dfr = dataframe_to_rows(data, header=header, index=index)

            if axis == 1:
                for c_idx, _column in enumerate(dfr, column):  # (Startcolumn)
                    for r_idx, value in enumerate(_column, row):  # (Startrow)
                        if (value is not None) and (str(value) != "nan"):
                            self.__change_xml(self.tree, r_idx, xl_name(c_idx), value)
                        else:
                            pass
            else:
                for r_idx, _row in enumerate(dfr, row):  # (Startrow)
                    for c_idx, value in enumerate(_row, column):  # (Startcolumn)
                        if (value is not None) and (str(value) != "nan"):
                            self.__change_xml(self.tree, r_idx, xl_name(c_idx), value)
                        else:
                            pass

        elif isinstance(data, self.check) or isinstance(data, str):
            self.__change_xml(self.tree, row, xl_name(column), data)

        self.__clean_formula()
        self.__write_xml()
        self.__write_strxml()

    def save(self, path: str = None) -> None:
        """
        Saves the converted Excel file to the specified path

        Args:
            path (str): The file path of the Excel workbook to read.

        Raises:
            ValueError: If the `path` argument is None.

        """

        if path is None:
            raise ValueError('Output path is missing')

        return shutil.copy(self.temp, path)
